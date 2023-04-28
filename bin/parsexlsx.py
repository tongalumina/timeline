#!/usr/bin/env python
import openpyxl
import os

# Load the Excel file and select the active worksheet
wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active

# Get the headers from the first row of the worksheet
headers = [cell.value for cell in ws[1]]

# Loop through the rows of the worksheet (excluding the header row)
for row in ws.iter_rows(min_row=2, values_only=True):
    # Create a dictionary of (key, value) pairs for this row
    row_dict = {}
    for i, cell_value in enumerate(row):
        if cell_value is None:
            row_dict[headers[i]] = "-"
        # Convert numeric cell values to integers
        elif isinstance(cell_value, (int, float)):
            row_dict[headers[i]] = int(cell_value)
        else:
            row_dict[headers[i]] = cell_value

    # Get the file name from the first column of the row
    if row[0] is None:
        file_name = "0000"
    else:
        file_name = str(int(row[0]))

    # Create a unique file name by appending a number if the file exists
    file_count = 1
    file_path = file_name + '_' + str(file_count) + '.txt'
    while os.path.exists(file_path):
        file_path = file_name + '_' + str(file_count) + '.txt'
        file_count += 1

    # Write the row data to the file
    with open(file_path, 'w') as f:
        for key, value in row_dict.items():
            f.write(">" + key + ':\n' + str(value) + '\n')
